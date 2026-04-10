import io
import pandas as pd
import numpy as np
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, Image, KeepTogether, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from analysis.chart_utils import (
        bar_chart, hbar_chart, grouped_bar, pie_chart,
        scatter_chart, funnel_chart, color_bar, hist_chart
    )
except ImportError:
    from chart_utils import (
        bar_chart, hbar_chart, grouped_bar, pie_chart,
        scatter_chart, funnel_chart, color_bar, hist_chart
    )

# ── Page geometry ─────────────────────────────────────────────────────────────
W    = A4[0] - 4*cm          # 481.89 pt = 17.000 cm
Wcm  = W / cm                # 17.0

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY  = colors.HexColor('#1A1A2E')
BLUE  = colors.HexColor('#378ADD')
PINK  = colors.HexColor('#D4537E')
TEAL  = colors.HexColor('#1D9E75')
GREEN = colors.HexColor('#639922')
GRNDK = colors.HexColor('#3B6D11')
AMBER = colors.HexColor('#BA7517')
RED   = colors.HexColor('#E24B4A')
LGREY = colors.HexColor('#F8F9FA')
MGREY = colors.HexColor('#E9ECEF')
WHITE = colors.white
DGREY = colors.HexColor('#6B7280')

# ── Paragraph styles ─────────────────────────────────────────────────────────
def PS(name, fs=9, bold=False, color=None, align=TA_LEFT, sp=3, lft=0):
    return ParagraphStyle(name,
        fontName='Helvetica-Bold' if bold else 'Helvetica',
        fontSize=fs, textColor=color or colors.HexColor('#2C2C2A'),
        spaceAfter=sp, spaceBefore=0, leading=fs+3,
        alignment=align, leftIndent=lft, wordWrap='LTR')

# Cell styles — all word-wrapping
CS      = PS('cell',    8, False)                          # normal cell
CS_B    = PS('cellb',   8, True)                           # bold cell
CS_C    = PS('cellc',   8, False, align=TA_CENTER)         # centered
CS_R    = PS('cellr',   8, False, align=TA_RIGHT)          # right
CS_SM   = PS('cellsm',  7, False)                          # small
CS_HDR  = PS('hdr',     8, True,  WHITE, TA_CENTER)        # header row
CS_W    = PS('cellw',   7.5, False)                        # wrap-heavy cells

def sp(n=8):   return Spacer(1, n)
def shorten(s, n): s=str(s); return s[:n]+'…' if len(s)>n else s
def P(text, style=None): return Paragraph(str(text), style or CS)
def PC(text): return Paragraph(str(text), CS_C)
def PW(text, maxlen=None):
    t = str(text)
    if maxlen: t = shorten(t, maxlen)
    return Paragraph(t, CS_W)

# ── Table builder — every cell wrapped in Paragraph ──────────────────────────
def mktbl(data, cw, hcol=BLUE, font_size=8):
    """
    Build a ReportLab table.
    - data: list of lists. First row = headers (plain strings).
            Data rows should contain Paragraph objects or plain strings.
    - cw:   list of column widths (must sum to W exactly).
    - hcol: header background colour.
    """
    assert abs(sum(cw) - W) < 1.0, f"Col widths sum {sum(cw):.2f} != W {W:.2f}"

    # Wrap header row
    hdr_style = ParagraphStyle('th', fontName='Helvetica-Bold',
                                fontSize=font_size, textColor=WHITE,
                                alignment=TA_CENTER, leading=font_size+3,
                                wordWrap='LTR')
    wrapped = []
    for ri, row in enumerate(data):
        new_row = []
        for ci, cell in enumerate(row):
            if ri == 0:
                new_row.append(Paragraph(str(cell), hdr_style))
            elif isinstance(cell, Paragraph):
                new_row.append(cell)
            else:
                # Auto-wrap plain strings
                st = CS_C if ci > 0 else CS
                new_row.append(Paragraph(str(cell), st))
        wrapped.append(new_row)

    ts = [
        ('BACKGROUND',    (0,0),(-1,0),  hcol),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 5),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('GRID',          (0,0),(-1,-1), 0.3, MGREY),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, LGREY]),
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
    ]
    t = Table(wrapped, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle(ts))
    return t

# ── Layout helpers ────────────────────────────────────────────────────────────
def banner(txt, col):
    t = Table([[Paragraph(txt, PS('bh',12,True,WHITE))]], colWidths=[W])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),col),
        ('TOPPADDING',(0,0),(-1,-1),8),
        ('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),12),
    ]))
    return t

def kpi_row(items):
    n  = len(items)
    cw = W / n
    cells = []
    for lbl, val, note, nc in items:
        inner = Table([
            [Paragraph(str(val),  PS('kv',13,True, NAVY,  TA_CENTER,0))],
            [Paragraph(str(lbl),  PS('kl', 8,False,DGREY, TA_CENTER,1))],
            [Paragraph(str(note), PS('kn', 8,False,nc,    TA_CENTER,0))],
        ], colWidths=[cw-6])
        inner.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),LGREY),
            ('BOX',(0,0),(-1,-1),0.5,MGREY),
            ('TOPPADDING',(0,0),(-1,-1),8),
            ('BOTTOMPADDING',(0,0),(-1,-1),8),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]))
        cells.append(inner)
    row = Table([cells], colWidths=[cw]*n)
    row.setStyle(TableStyle([
        ('LEFTPADDING',(0,0),(-1,-1),3),
        ('RIGHTPADDING',(0,0),(-1,-1),3),
    ]))
    return row

def two_charts(img1, img2):
    hw = W/2
    t  = Table([[img1, img2]], colWidths=[hw, hw])
    t.setStyle(TableStyle([
        ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
    ]))
    return t

def header_tbl(title, subtitle, period, analyst):
    rows = [
        [Paragraph(title,    PS('ht',20,True, WHITE,TA_CENTER,4))],
        [Paragraph(subtitle, PS('hs',10,False,colors.HexColor('#9CA3AF'),TA_CENTER,4))],
        [Paragraph(f"Period: {period}  ·  Analyst: {analyst}  ·  {datetime.now().strftime('%d %b %Y %H:%M')}",
                   PS('hd',9,False,colors.HexColor('#9CA3AF'),TA_CENTER,0))],
    ]
    t = Table(rows, colWidths=[W])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),NAVY),
        ('TOPPADDING',(0,0),(-1,-1),16),('BOTTOMPADDING',(0,0),(-1,-1),16),
        ('LEFTPADDING',(0,0),(-1,-1),12)]))
    return t

def footer_tbl(analyst, period, channel):
    t = Table([[Paragraph(
        f"iDealz {channel} Report  ·  {analyst}  ·  {period}  ·  Confidential",
        PS('ft',8,False,colors.HexColor('#9CA3AF'),TA_CENTER))]],colWidths=[W])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),NAVY),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
    return t

def alert_table(alerts, filter_key=None):
    filtered = [a for a in alerts if not filter_key or
                filter_key.lower() in a.get('title','').lower()] if alerts else []
    if not filtered: return None
    rows = [['#','Alert','Action Required']]
    for i,a in enumerate(filtered,1):
        rows.append([str(i), PW(a['title'],40), PW(a['msg'],80)])
    # col widths: 0.6 + 5 + rest
    cw = [0.6*cm, 5*cm, W-5.6*cm]
    return mktbl(rows, cw, RED)


# ═════════════════════════════════════════════════════════════════════════════
# META PDF REPORT
# ═════════════════════════════════════════════════════════════════════════════
def generate_meta_pdf(meta_df, analyst, period, alerts=None, meta_prev=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=1.5*cm, bottomMargin=2*cm)
    story = []

    story += [header_tbl("iDealz Meta Ads Report",
                          "Facebook and Instagram Ad Performance Analysis",
                          period, analyst), sp(12)]

    al = alert_table(alerts)
    if al: story += [banner("⚠  Alerts", RED), sp(6), al, sp(12)]

    # ── KPIs ──
    ts  = meta_df['Amount spent (LKR)'].sum()
    tr  = int(meta_df['Results'].sum())
    cpr = ts/tr if tr>0 else 0
    ctr = meta_df['CTR (link click-through rate)'].mean()
    cpm = meta_df['CPM (cost per 1,000 impressions)'].mean()
    cpc = meta_df['CPC (cost per link click)'].mean()
    frq = meta_df['Frequency'].mean()
    impr= int(meta_df['Impressions'].sum())
    rch = int(meta_df['Reach'].sum())
    clk = int(meta_df['Link clicks'].sum())

    story += [banner("📊  Performance Overview", BLUE), sp(8)]
    story.append(kpi_row([
        ('Total Spend',   f"Rs {ts:,.0f}", '—', DGREY),
        ('Total Results', f"{tr:,}",        '—', DGREY),
        ('Cost / Result', f"Rs {cpr:,.2f}", '⚠ High' if cpr>25 else '✅ Good',
         RED if cpr>25 else GREEN),
        ('Avg CTR',       f"{ctr:.2f}%",    '✅ Good' if ctr>=1.5 else '⚠ Low',
         GREEN if ctr>=1.5 else AMBER),
        ('Avg CPM',       f"Rs {cpm:,.0f}", '—', DGREY),
        ('Avg Frequency', f"{frq:.2f}",     '⚠ High' if frq>3 else '✅ OK',
         RED if frq>3 else GREEN),
    ]))
    story.append(sp(6))
    # 4-col KPI bar — equal widths
    kpi2 = [['Total Impressions','Total Reach','Total Clicks','Avg CPC']]
    kpi2.append([f"{impr:,}", f"{rch:,}", f"{clk:,}", f"Rs {cpc:,.2f}"])
    story += [mktbl(kpi2, [W/4]*4, BLUE), sp(10)]

    if meta_prev is not None:
        pv_ts=meta_prev['Amount spent (LKR)'].sum()
        pv_tr=int(meta_prev['Results'].sum())
        pv_rch=int(meta_prev['Reach'].sum())
        prev_rows = [['Metric','This Period','Previous Period','Change']]
        for m,cv,pv in [('Spend (LKR)',ts,pv_ts),('Results',tr,pv_tr),('Reach',rch,pv_rch)]:
            fv = f"Rs {cv:,.0f}" if 'Spend' in m else f"{int(cv):,}"
            fp = f"Rs {pv:,.0f}" if 'Spend' in m else f"{int(pv):,}"
            chg = f"{(cv-pv)/pv*100:+.1f}%" if pv>0 else '—'
            prev_rows.append([m, fv, fp, chg])
        story += [mktbl(prev_rows,[4*cm,4*cm,4*cm,W-12*cm],BLUE), sp(10)]

    # ── Platform FB vs IG ──
    story += [sp(4), banner("📱  Facebook vs Instagram", BLUE), sp(8)]
    if 'Platform' in meta_df.columns:
        plat = meta_df.groupby('Platform').agg(
            Spend=('Amount spent (LKR)','sum'), Results=('Results','sum'),
            Avg_CPM=('CPM (cost per 1,000 impressions)','mean'),
            Avg_CPC=('CPC (cost per link click)','mean'),
            Avg_CTR=('CTR (link click-through rate)','mean'),
            Avg_Freq=('Frequency','mean'),
        ).round(2).reset_index()
        plat['CPR'] = (plat['Spend']/plat['Results'].replace(0,1)).round(2)
        pcolors = ['#378ADD' if 'face' in str(p).lower() else '#D4537E'
                   for p in plat['Platform']]

        c1 = bar_chart(plat['Platform'].tolist(), plat['Results'].tolist(),
                       'Results by Platform', color=pcolors, W=Wcm/2-0.5, H=6)
        c2 = bar_chart(plat['Platform'].tolist(), plat['CPR'].tolist(),
                       'Cost Per Result (Rs)', color=pcolors,
                       fmt='Rs {:.0f}', W=Wcm/2-0.5, H=6)
        c3 = bar_chart(plat['Platform'].tolist(), plat['Avg_CTR'].tolist(),
                       'Avg CTR (%)', color=pcolors, fmt='{:.2f}%', W=Wcm/2-0.5, H=6)
        c4 = bar_chart(plat['Platform'].tolist(), plat['Avg_CPM'].tolist(),
                       'Avg CPM (Rs)', color=pcolors, fmt='Rs {:.0f}', W=Wcm/2-0.5, H=6)

        # Platform table — 8 cols, fixed widths totalling W exactly
        # [3.0, 2.8, 1.8, 2.5, 1.8, 2.2, 2.2, 0.7] = 17.0cm
        pt_cw = [2.8*cm,2.8*cm,1.8*cm,2.5*cm,1.8*cm,2.2*cm,2.2*cm,W-16.1*cm]
        pt_rows = [['Platform','Spend (LKR)','Results','Cost/Result','CTR','CPM','CPC','Freq']]
        for _,r in plat.iterrows():
            pt_rows.append([
                PW(r['Platform'].title(), 12),
                PC(f"Rs {r['Spend']:,.0f}"),
                PC(str(int(r['Results']))),
                PC(f"Rs {r['CPR']:,.2f}"),
                PC(f"{r['Avg_CTR']:.2f}%"),
                PC(f"Rs {r['Avg_CPM']:,.0f}"),
                PC(f"Rs {r['Avg_CPC']:,.0f}"),
                PC(f"{r['Avg_Freq']:.2f}"),
            ])
        story += [two_charts(c1,c2), sp(4), two_charts(c3,c4), sp(6),
                  mktbl(pt_rows, pt_cw, BLUE), sp(12)]

    # ── Ad Set Deep Dive ──
    story += [banner("📂  Ad Set Analysis — CPM · CPC · CTR", BLUE), sp(8)]
    if 'Ad set name' in meta_df.columns:
        adset = meta_df.groupby('Ad set name').agg(
            Spend=('Amount spent (LKR)','sum'), Results=('Results','sum'),
            Avg_CPM=('CPM (cost per 1,000 impressions)','mean'),
            Avg_CPC=('CPC (cost per link click)','mean'),
            Avg_CTR=('CTR (link click-through rate)','mean'),
            Avg_Freq=('Frequency','mean'),
        ).round(2).reset_index()
        adset['CPR'] = (adset['Spend']/adset['Results'].replace(0,1)).round(2)
        adset_s = adset.sort_values('Avg_CTR', ascending=False)
        n_sets  = len(adset)
        h_ht    = max(7, n_sets*0.55)

        # Full-width horizontal charts (auto-triggered when >10 items in color_bar)
        c1 = color_bar(adset_s['Ad set name'].tolist(),
                       adset_s['Avg_CTR'].tolist(),
                       'CTR by Ad Set (%)',
                       cmap_name='Greens', fmt='{:.2f}%', W=Wcm, H=h_ht)
        c2 = color_bar(adset.sort_values('Avg_CPM')['Ad set name'].tolist(),
                       adset.sort_values('Avg_CPM')['Avg_CPM'].tolist(),
                       'CPM by Ad Set (Rs) — lower = better',
                       cmap_name='RdYlGn_r', fmt='Rs {:.0f}', W=Wcm, H=h_ht)
        c3 = color_bar(adset.sort_values('Avg_CPC')['Ad set name'].tolist(),
                       adset.sort_values('Avg_CPC')['Avg_CPC'].tolist(),
                       'CPC by Ad Set (Rs) — lower = better',
                       cmap_name='RdYlGn_r', fmt='Rs {:.0f}', W=Wcm, H=h_ht)
        c4 = color_bar(adset.sort_values('Results',ascending=False)['Ad set name'].tolist(),
                       adset.sort_values('Results',ascending=False)['Results'].tolist(),
                       'Results by Ad Set', cmap_name='Blues', W=Wcm, H=h_ht)
        story += [c1, sp(6), c2, sp(6), c3, sp(6), c4, sp(6)]

        # Ad set table — 8 cols
        # [3.5, 2.5, 1.8, 2.5, 1.8, 2.2, 2.2, 0.5] = 17.0cm
        as_cw = [4.0*cm,2.3*cm,1.5*cm,2.3*cm,1.8*cm,2.2*cm,2.2*cm,W-16.3*cm]
        as_rows = [['Ad Set','Spend','Results','CPR','CTR','CPM','CPC','Freq']]
        for _,r in adset.sort_values('Results',ascending=False).iterrows():
            as_rows.append([
                PW(r['Ad set name'], 18),
                PC(f"Rs {r['Spend']:,.0f}"),
                PC(str(int(r['Results']))),
                PC(f"Rs {r['CPR']:,.2f}"),
                PC(f"{r['Avg_CTR']:.2f}%"),
                PC(f"Rs {r['Avg_CPM']:,.0f}"),
                PC(f"Rs {r['Avg_CPC']:,.0f}"),
                PC(f"{r['Avg_Freq']:.2f}"),
            ])
        story += [mktbl(as_rows, as_cw, BLUE), sp(12)]

    # ── Campaign Pacing ──
    story += [PageBreak(), banner("💰  Campaign Pacing vs Budget", BLUE), sp(8)]
    if 'Ad set name' in meta_df.columns:
        pacing = meta_df.groupby('Ad set name').agg(
            Spend=('Amount spent (LKR)','sum'), Results=('Results','sum')
        ).round(2).reset_index()
        avg_sp = pacing['Spend'].mean()
        pacing['vs_avg'] = ((pacing['Spend']-avg_sp)/avg_sp*100).round(1)
        pacing['Status'] = pacing['vs_avg'].apply(
            lambda x:'🔴 Overspending' if x>30 else(
                     '🟡 Slightly over' if x>10 else(
                     '🟢 On track'      if x>-10 else '⚠ Underspending')))
        pc_s  = pacing.sort_values('Spend', ascending=False)
        p_ht  = max(7, len(pc_s)*0.55)
        chart = color_bar(pc_s['Ad set name'].tolist(), pc_s['Spend'].tolist(),
                          'Spend by Ad Set (Rs)', cmap_name='RdYlGn_r',
                          fmt='Rs {:.0f}', W=Wcm, H=p_ht)
        story += [chart, sp(6)]

        # Pacing table — 5 cols [3.5, 2.8, 1.8, 2.2, rest]
        pc_cw = [4.5*cm,2.5*cm,1.5*cm,2.0*cm,W-10.5*cm]
        pc_rows = [['Ad Set','Spend (LKR)','Results','vs Avg','Status']]
        for _,r in pc_s.iterrows():
            pc_rows.append([
                PW(r['Ad set name'], 18),
                PC(f"Rs {r['Spend']:,.0f}"),
                PC(str(int(r['Results']))),
                PC(f"{r['vs_avg']:+.1f}%"),
                P(r['Status']),
            ])
        story += [mktbl(pc_rows, pc_cw, BLUE), sp(12)]

    # ── Top & Bottom Ads ──
    story += [banner("🏆  Top and Bottom Performing Ads", BLUE), sp(8)]
    active = meta_df[meta_df['Results']>0].copy()
    if len(active)>0:
        top10 = active.sort_values('Cost per result').head(10)
        bot10 = active.sort_values('Cost per result', ascending=False).head(10)

        c1 = hbar_chart(top10['Ad name'].str[:25].tolist(),
                        top10['Cost per result'].tolist(),
                        'Top 10 — Lowest Cost per Result (Rs)',
                        colormap='Greens', fmt='Rs {:.0f}',
                        W=Wcm/2-0.5, H=max(7, len(top10)*0.7))
        c2 = hbar_chart(bot10['Ad name'].str[:25].tolist(),
                        bot10['Cost per result'].tolist(),
                        'Bottom 10 — Highest Cost per Result (Rs)',
                        colormap='Reds', fmt='Rs {:.0f}',
                        W=Wcm/2-0.5, H=max(7, len(bot10)*0.7))
        story += [two_charts(c1,c2), sp(6)]

        # Scatter chart
        pcolors_s = ['#378ADD' if 'face' in str(p).lower() else '#D4537E'
                     for p in active['Platform']]
        c_sc = scatter_chart(
            active['Amount spent (LKR)'].tolist(),
            active['CTR (link click-through rate)'].tolist(),
            active['Ad name'].tolist(),
            'Spend vs CTR per Ad  (bubble size = results)',
            hline=1.5, hline_label='1.5% CTR benchmark',
            colors_list=pcolors_s,
            sizes=active['Results'].tolist(),
            W=Wcm, H=7,
            xlabel='Amount Spent (LKR)', ylabel='CTR (%)')
        story += [c_sc, sp(6)]

        # Tables — 6 cols [4.5, 2.0, 1.8, 2.5, 2.5, rest]
        ads_cw = [4.5*cm,2.0*cm,1.8*cm,2.5*cm,2.5*cm,W-13.3*cm]
        story.append(P("Top 10 Ads — Lowest Cost per Result:", PS('h3',9,True,NAVY,sp=4)))
        t5r = [['Ad Name','Platform','Results','Cost/Result','Spend','CTR']]
        for _,r in top10.iterrows():
            t5r.append([
                PW(r['Ad name'], 20),
                PC(r['Platform'].title()),
                PC(str(int(r['Results']))),
                PC(f"Rs {r['Cost per result']:,.2f}"),
                PC(f"Rs {r['Amount spent (LKR)']:,.0f}"),
                PC(f"{r['CTR (link click-through rate)']:.2f}%"),
            ])
        story += [mktbl(t5r, ads_cw, BLUE), sp(8)]

        story.append(P("Bottom 10 Ads — Highest Cost per Result (review for pausing):",
                       PS('h3',9,True,RED,sp=4)))
        b5r = [['Ad Name','Platform','Results','Cost/Result','Spend','CTR']]
        for _,r in bot10.iterrows():
            b5r.append([
                PW(r['Ad name'], 20),
                PC(r['Platform'].title()),
                PC(str(int(r['Results']))),
                PC(f"Rs {r['Cost per result']:,.2f}"),
                PC(f"Rs {r['Amount spent (LKR)']:,.0f}"),
                PC(f"{r['CTR (link click-through rate)']:.2f}%"),
            ])
        story += [mktbl(b5r, ads_cw, colors.HexColor('#A03060')), sp(12)]

    # ── Frequency & Fatigue ──
    story += [banner("🔄  Frequency and Audience Fatigue", BLUE), sp(8)]
    fatigued = meta_df[meta_df['Frequency']>3]
    if len(fatigued)>0:
        story.append(P(f"⚠  {len(fatigued)} ads have frequency above 3.0 — audience needs refreshing.",
                       PS('fa',9,False,RED,sp=4)))
        fa_cw = [5.0*cm,2.0*cm,2.0*cm,3.0*cm,W-12*cm]
        fa_rows = [['Ad Name','Platform','Frequency','Spend','Results']]
        for _,r in fatigued.sort_values('Frequency',ascending=False).iterrows():
            fa_rows.append([PW(r['Ad name'], 22), PC(r['Platform'].title()),
                PC(f"{r['Frequency']:.2f}"), PC(f"Rs {r['Amount spent (LKR)']:,.0f}"),
                PC(str(int(r['Results'])))])
        story += [mktbl(fa_rows, fa_cw, RED), sp(8)]
    else:
        story.append(P("✅  No fatigued audiences — all frequency below 3.0",
                       PS('ok',9,False,GREEN,sp=6)))

    chart_h = hist_chart(meta_df['Frequency'].dropna().tolist(),
                         'Frequency Distribution', color='#378ADD',
                         vline=3, vline_label='Fatigue threshold (3.0)',
                         xlabel='Frequency', ylabel='Number of Ads',
                         W=Wcm, H=6, bins=15)
    story += [chart_h, sp(8)]

    zero = meta_df[(meta_df['Results']==0)&(meta_df['Amount spent (LKR)']>100)]
    if len(zero)>0:
        story.append(P(f"⚠  {len(zero)} zero-result ads spending budget — consider pausing.",
                       PS('zr',9,False,RED,sp=4)))
        z_cw  = [5.5*cm,2.0*cm,3.0*cm,W-10.5*cm]
        z_rows= [['Ad Name','Platform','Spend','Impressions']]
        for _,r in zero.iterrows():
            z_rows.append([PW(r['Ad name'], 24), PC(r['Platform'].title()),
                PC(f"Rs {r['Amount spent (LKR)']:,.0f}"), PC(f"{int(r['Impressions']):,}")])
        story += [mktbl(z_rows, z_cw, RED), sp(8)]

    # ── Placement ──
    if 'Placement' in meta_df.columns:
        story += [banner("📍  Placement Breakdown", BLUE), sp(8)]
        place = meta_df.groupby('Placement').agg(
            Spend=('Amount spent (LKR)','sum'), Results=('Results','sum'),
            Avg_CTR=('CTR (link click-through rate)','mean'),
            Avg_CPM=('CPM (cost per 1,000 impressions)','mean'),
        ).round(2).reset_index().sort_values('Results',ascending=False)
        place['CPR'] = (place['Spend']/place['Results'].replace(0,1)).round(2)

        c1 = pie_chart(place['Placement'].tolist(), place['Spend'].tolist(),
                       'Spend by Placement', W=Wcm/2-0.5, H=7)
        c2 = bar_chart(place['Placement'].tolist(), place['Avg_CTR'].tolist(),
                       'CTR by Placement (%)', color='#639922',
                       fmt='{:.2f}%', W=Wcm/2-0.5, H=7, rotate=20)
        story += [two_charts(c1,c2), sp(6)]

        # Placement table — 6 cols [4.0, 2.8, 2.0, 2.5, 2.0, rest]
        pl_cw = [4.0*cm,2.8*cm,2.0*cm,2.5*cm,2.0*cm,W-13.3*cm]
        pl_rows = [['Placement','Spend','Results','CPR','CTR','CPM']]
        for _,r in place.iterrows():
            pl_rows.append([
                PW(r['Placement'], 18),
                PC(f"Rs {r['Spend']:,.0f}"),
                PC(str(int(r['Results']))),
                PC(f"Rs {r['CPR']:,.2f}"),
                PC(f"{r['Avg_CTR']:.2f}%"),
                PC(f"Rs {r['Avg_CPM']:,.0f}"),
            ])
        story += [mktbl(pl_rows, pl_cw, BLUE), sp(12)]

    # ── Monthly Wrap ──
    story += [banner("📈  Monthly Performance Summary", BLUE), sp(8)]
    story.append(kpi_row([
        ('Total Spend',       f"Rs {ts:,.0f}",  '—', DGREY),
        ('Total Results',     f"{tr:,}",          '—', DGREY),
        ('Total Impressions', f"{impr:,}",        '—', DGREY),
        ('Total Reach',       f"{rch:,}",         '—', DGREY),
        ('Total Clicks',      f"{clk:,}",         '—', DGREY),
    ]))
    story.append(sp(8))
    if meta_prev is None:
        story.append(P("Upload previous period Meta export in the sidebar to see month-over-month comparison.",
                       PS('mo',9,False,AMBER,sp=6)))

    story += [sp(12), footer_tbl(analyst, period, "Meta Ads")]
    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════════════════════
# TIKTOK PDF REPORT  (already working — minimal wrapper)
# ═════════════════════════════════════════════════════════════════════════════
def generate_tiktok_pdf(df, analyst, period, alerts=None, prev_df=None):
    from analysis.report_generator import _build_tiktok
    return _build_tiktok(df, analyst, period, alerts, prev_df)


# ═════════════════════════════════════════════════════════════════════════════
# WEBSITE PDF REPORT
# ═════════════════════════════════════════════════════════════════════════════
def generate_website_pdf(ga4_bundle, analyst, period, alerts=None, gsc_bundle=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=1.5*cm, bottomMargin=2*cm)
    story = []

    traffic_ch  = ga4_bundle.get('traffic_channel')
    traffic_src = ga4_bundle.get('traffic_source')
    users_df    = ga4_bundle.get('users')
    pages_df    = ga4_bundle.get('pages')
    events_df   = ga4_bundle.get('events')
    gsc_q       = gsc_bundle.get('queries') if gsc_bundle else None

    story += [header_tbl("iDealz Website Report",
                          "GA4 Traffic, Engagement and SEO Analysis",
                          period, analyst), sp(12)]

    al = alert_table(alerts)
    if al: story += [banner("⚠  Alerts", RED), sp(6), al, sp(12)]

    # ── Traffic Overview ──
    if traffic_ch is not None:
        story += [banner("📊  Traffic Overview", GRNDK), sp(8)]
        col0 = traffic_ch.columns[0]
        ts_w = int(traffic_ch['Sessions'].sum())
        es_w = int(traffic_ch['Engaged sessions'].sum()) if 'Engaged sessions' in traffic_ch.columns else 0
        er_w = traffic_ch['Engagement rate'].mean()*100   if 'Engagement rate' in traffic_ch.columns else 0
        ad_w = traffic_ch['Average engagement time per session'].mean() if 'Average engagement time per session' in traffic_ch.columns else 0
        ev_w = int(traffic_ch['Event count'].sum()) if 'Event count' in traffic_ch.columns else 0

        story.append(kpi_row([
            ('Total Sessions',   f"{ts_w:,}",    '—', DGREY),
            ('Engaged Sessions', f"{es_w:,}",    '—', DGREY),
            ('Engagement Rate',  f"{er_w:.1f}%", '✅ Good' if er_w>50 else '⚠ Low',
             GREEN if er_w>50 else AMBER),
            ('Avg Duration',     f"{ad_w:.0f}s",  '✅ Good' if ad_w>45 else '⚠ Low',
             GREEN if ad_w>45 else AMBER),
            ('Total Events',     f"{ev_w:,}",    '—', DGREY),
        ]))
        story.append(sp(8))

        channels = traffic_ch[col0].astype(str).str.lower().tolist()
        if not any('paid' in c for c in channels):
            story.append(P(
                "No Paid Social channel in GA4 — Meta and TikTok ad spend is not tracked. "
                "Add UTM parameters to all ad URLs:  "
                "utm_source=facebook  utm_medium=cpc  (Meta) and "
                "utm_source=tiktok  utm_medium=cpc  (TikTok).",
                PS('utmw',9,False,RED,sp=8)))

        c1 = bar_chart(traffic_ch[col0].tolist(), traffic_ch['Sessions'].tolist(),
                       'Sessions by Channel', color='#3B6D11', W=Wcm/2-0.5, H=6, rotate=15)
        c2 = pie_chart(traffic_ch[col0].tolist(), traffic_ch['Sessions'].tolist(),
                       'Session Share', W=Wcm/2-0.5, H=6)
        story += [two_charts(c1,c2), sp(6)]

        if 'Engagement rate' in traffic_ch.columns:
            traffic_ch['Eng%'] = (traffic_ch['Engagement rate']*100).round(1)
            c3 = color_bar(traffic_ch[col0].tolist(), traffic_ch['Eng%'].tolist(),
                           'Engagement Rate by Channel (%)',
                           cmap_name='RdYlGn', fmt='{:.1f}%', W=Wcm, H=5)
            story += [c3, sp(4)]
        story.append(P("Benchmark:  below 40% = Low   40–60% = Average   above 60% = Good",
                       PS('bm',8,False,AMBER,sp=6)))

        # Traffic table — 6 cols [4.5, 2.0, 2.0, 2.0, 2.5, rest]
        tc_cw = [4.5*cm,2.0*cm,2.0*cm,2.0*cm,2.5*cm,W-13*cm]
        tc_rows = [['Channel','Sessions','Engaged','Eng. Rate','Avg Duration','Events']]
        for _,r in traffic_ch.iterrows():
            tc_rows.append([
                PW(r.iloc[0], 20),
                PC(f"{int(r['Sessions']):,}"),
                PC(f"{int(r['Engaged sessions']):,}" if 'Engaged sessions' in traffic_ch.columns else '—'),
                PC(f"{r['Engagement rate']*100:.1f}%"     if 'Engagement rate' in traffic_ch.columns else '—'),
                PC(f"{r['Average engagement time per session']:.0f}s" if 'Average engagement time per session' in traffic_ch.columns else '—'),
                PC(f"{int(r['Event count']):,}" if 'Event count' in traffic_ch.columns else '—'),
            ])
        story += [mktbl(tc_rows, tc_cw, GRNDK), sp(12)]

    # ── Source / Medium ──
    if traffic_src is not None:
        story += [banner("🔍  Traffic Sources — Source and Medium", GRNDK), sp(8)]
        src_col = traffic_src.columns[0]
        top15   = traffic_src.sort_values('Sessions',ascending=False).head(15)

        c1 = hbar_chart(top15[src_col].str[:35].tolist(), top15['Sessions'].tolist(),
                        'Top 15 Sources by Sessions', colormap='Greens',
                        W=Wcm, H=max(8, len(top15)*0.7))
        story += [c1, sp(6)]

        # Source table — 4 cols [5.5, 2.0, 2.0, rest]
        sc_cw = [5.5*cm,2.0*cm,2.0*cm,W-9.5*cm]
        sc_rows = [['Source / Medium','Sessions','Eng. Rate','Avg Duration']]
        for _,r in top15.iterrows():
            sc_rows.append([
                PW(r.iloc[0], 24),
                PC(f"{int(r['Sessions']):,}"),
                PC(f"{r['Engagement rate']*100:.1f}%" if 'Engagement rate' in traffic_src.columns else '—'),
                PC(f"{r['Average engagement time per session']:.0f}s" if 'Average engagement time per session' in traffic_src.columns else '—'),
            ])
        story += [mktbl(sc_rows, sc_cw, GRNDK), sp(6)]
        story.append(P(
            "UTM Fix — add to all ad destination URLs:   "
            "Meta: utm_source=facebook  utm_medium=cpc   "
            "TikTok: utm_source=tiktok  utm_medium=cpc",
            PS('utmf',8,False,AMBER,sp=8)))
        story.append(sp(6))

    # ── Top Pages ──
    if pages_df is not None:
        story += [banner("📄  Top Pages — Engagement Analysis", GRNDK), sp(8)]
        pg_col = pages_df.columns[0]
        total_views = int(pages_df['Views'].sum()) if 'Views' in pages_df.columns else 0
        story.append(P(f"Total page views: {total_views:,}   |   Unique pages tracked: {len(pages_df):,}",
                       PS('pgsum',9,False,NAVY,sp=6)))

        top15_pg = pages_df.head(15)
        c1 = hbar_chart(top15_pg[pg_col].str[-35:].tolist(), top15_pg['Views'].tolist(),
                        'Top 15 Pages by Views', colormap='Greens',
                        W=Wcm, H=max(8,len(top15_pg)*0.7))
        story += [c1, sp(6)]

        if 'Average engagement time per active user' in pages_df.columns:
            top_eng = pages_df.nlargest(15,'Average engagement time per active user')
            c2 = hbar_chart(top_eng[pg_col].str[-35:].tolist(),
                            top_eng['Average engagement time per active user'].tolist(),
                            'Avg Engagement Time by Page (seconds)',
                            colormap='Blues', fmt='{:.0f}s',
                            W=Wcm, H=max(8,len(top_eng)*0.7))
            story += [c2, sp(6)]

        # Pages table — 5 cols [6.5, 1.5, 2.0, 2.5, rest]
        pg_cw = [6.5*cm,1.5*cm,2.0*cm,2.5*cm,W-12.5*cm]
        pg_rows = [['Page Path','Views','Active Users','Avg Eng. Time','Events']]
        for _,r in top15_pg.iterrows():
            pg_rows.append([
                PW(r.iloc[0], 29),
                PC(f"{int(r['Views']):,}"),
                PC(f"{int(r['Active users']):,}" if 'Active users' in pages_df.columns else '—'),
                PC(f"{r['Average engagement time per active user']:.0f}s" if 'Average engagement time per active user' in pages_df.columns else '—'),
                PC(f"{int(r['Event count']):,}" if 'Event count' in pages_df.columns else '—'),
            ])
        story += [mktbl(pg_rows, pg_cw, GRNDK), sp(12)]

    # ── Users ──
    if users_df is not None:
        story += [banner("👥  User Acquisition", GRNDK), sp(8)]
        total_u = int(users_df['Total users'].sum())    if 'Total users'    in users_df.columns else 0
        new_u   = int(users_df['New users'].sum())      if 'New users'      in users_df.columns else 0
        ret_u   = int(users_df['Returning users'].sum())if 'Returning users' in users_df.columns else 0
        story.append(kpi_row([
            ('Total Users',     f"{total_u:,}", '—', DGREY),
            ('New Users',       f"{new_u:,}",
             f"{new_u/total_u*100:.0f}% of total" if total_u>0 else '—', BLUE),
            ('Returning Users', f"{ret_u:,}",
             f"{ret_u/total_u*100:.0f}% returning" if total_u>0 else '—', TEAL),
        ]))
        story.append(sp(8))
        usr_col = users_df.columns[0]
        if 'New users' in users_df.columns and 'Returning users' in users_df.columns:
            c1 = bar_chart(users_df[usr_col].tolist(), users_df['New users'].tolist(),
                           'New Users by Channel', color='#378ADD',
                           W=Wcm/2-0.5, H=6, rotate=15)
            c2 = bar_chart(users_df[usr_col].tolist(), users_df['Returning users'].tolist(),
                           'Returning Users by Channel', color='#1D9E75',
                           W=Wcm/2-0.5, H=6, rotate=15)
            story += [two_charts(c1,c2), sp(6)]
        # Users table — 4 cols [5, 3, 3, rest]
        u_cw = [5*cm,3*cm,3*cm,W-11*cm]
        u_rows = [['Channel','Total Users','New Users','Returning Users']]
        for _,r in users_df.iterrows():
            u_rows.append([
                PW(r.iloc[0], 22),
                PC(f"{int(r['Total users']):,}"    if 'Total users'    in users_df.columns else '—'),
                PC(f"{int(r['New users']):,}"       if 'New users'      in users_df.columns else '—'),
                PC(f"{int(r['Returning users']):,}" if 'Returning users' in users_df.columns else '—'),
            ])
        story += [mktbl(u_rows, u_cw, GRNDK), sp(12)]

    # ── Events ──
    if events_df is not None:
        story += [banner("⚡  Events Tracked", GRNDK), sp(8)]
        evt_col = events_df.columns[0]
        total_e = int(events_df['Event count'].sum()) if 'Event count' in events_df.columns else 0
        story.append(P(f"Total events: {total_e:,}   |   Event types tracked: {len(events_df)}",
                       PS('esum',9,False,NAVY,sp=4)))
        if len(events_df)<=4:
            story.append(P(
                "Only 4 basic events tracked. WhatsApp click and purchase events are not set up yet. "
                "To track WhatsApp clicks: add gtag('event','whatsapp_click') when the button is clicked, "
                "then mark it as a Key Event in GA4 Admin.",
                PS('evw',9,False,AMBER,sp=6)))
        if 'Event count' in events_df.columns:
            c1 = pie_chart(events_df[evt_col].tolist(), events_df['Event count'].tolist(),
                           'Event Distribution', W=Wcm/2-0.5, H=7)
            c2 = bar_chart(events_df[evt_col].tolist(), events_df['Event count'].tolist(),
                           'Event Count', color='#3B6D11', W=Wcm/2-0.5, H=7)
            story += [two_charts(c1,c2), sp(6)]
        # Events table — 4 cols [4, 3, 3, rest]
        ev_cw = [4*cm,3*cm,3*cm,W-10*cm]
        ev_rows = [['Event Name','Count','Total Users','Per User']]
        for _,r in events_df.iterrows():
            ev_rows.append([
                PW(r.iloc[0], 20),
                PC(f"{int(r['Event count']):,}"),
                PC(f"{int(r['Total users']):,}"),
                PC(f"{r['Event count per active user']:.2f}" if 'Event count per active user' in events_df.columns else '—'),
            ])
        story += [mktbl(ev_rows, ev_cw, GRNDK), sp(12)]

    # ── SEO ──
    story += [banner("🔎  SEO Performance", GRNDK), sp(8)]
    if gsc_q is not None:
        total_gc = int(gsc_q['Clicks'].sum()) if 'Clicks' in gsc_q.columns else 0
        total_gi = int(gsc_q['Impressions'].sum()) if 'Impressions' in gsc_q.columns else 0
        avg_pos  = gsc_q['Position'].mean() if 'Position' in gsc_q.columns else 0
        story.append(kpi_row([
            ('Search Clicks',      f"{total_gc:,}",  '—', DGREY),
            ('Search Impressions', f"{total_gi:,}",  '—', DGREY),
            ('Avg Position',       f"{avg_pos:.1f}",
             '✅ Page 1' if avg_pos<10 else '⚠ Below page 1',
             GREEN if avg_pos<10 else AMBER),
        ]))
        story.append(sp(8))
        q_col = gsc_q.columns[0]
        if 'Clicks' in gsc_q.columns:
            top20 = gsc_q.sort_values('Clicks',ascending=False).head(20)
            c1 = hbar_chart(top20[q_col].str[:35].tolist(), top20['Clicks'].tolist(),
                            'Top 20 Queries by Clicks', colormap='Greens',
                            W=Wcm, H=max(8,len(top20)*0.7))
            story += [c1, sp(6)]
            gq_cw = [5.5*cm,2.0*cm,2.5*cm,2.0*cm,W-12*cm]
            gq_rows = [['Query','Clicks','Impressions','CTR','Position']]
            for _,r in top20.iterrows():
                gq_rows.append([
                    PW(r.iloc[0], 24),
                    PC(f"{int(r['Clicks']):,}"),
                    PC(f"{int(r['Impressions']):,}"),
                    PC(f"{r['CTR']*100:.1f}%"   if 'CTR'      in gsc_q.columns else '—'),
                    PC(f"{r['Position']:.1f}"   if 'Position' in gsc_q.columns else '—'),
                ])
            story += [mktbl(gq_rows, gq_cw, GRNDK), sp(12)]
    else:
        story.append(P(
            "Upload Google Search Console Queries CSV to see keyword performance. "
            "Export: Search Console → Performance → Search Results → Export CSV.",
            PS('seon',9,False,AMBER,sp=8)))
        story.append(sp(8))

    story += [sp(12), footer_tbl(analyst, period, "Website")]
    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════════════════════
# TIKTOK — internal builder (called by generate_tiktok_pdf)
# ═════════════════════════════════════════════════════════════════════════════
def _build_tiktok(df, analyst, period, alerts=None, prev_df=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=1.5*cm, bottomMargin=2*cm)
    story = []

    story += [header_tbl("iDealz TikTok Ads Report",
                          "Video Performance and Engagement Analysis",
                          period, analyst), sp(12)]

    al = alert_table(alerts, filter_key='tiktok')
    if al: story += [banner("⚠  Alerts", RED), sp(6), al, sp(12)]

    ts  = df['Cost'].sum()
    tv  = int(df['Video views'].sum())
    aw  = df['Average play time per video view'].mean()
    ac  = df['100% video view rate'].mean()*100
    td  = int(df['Clicks (destination)'].sum())
    ti  = int(df['Impressions'].sum())
    tr  = int(df['Reach'].sum())
    tf  = df['Frequency'].mean()
    tall= int(df['Clicks (all)'].sum())

    story += [banner("📊  Performance Overview", PINK), sp(8)]
    story.append(kpi_row([
        ('Total Spend',     f"${ts:,.2f}", '—', DGREY),
        ('Video Views',     f"{tv:,}",      '—', DGREY),
        ('Avg Watch Time',  f"{aw:.1f}s",   '✅ Good' if aw>=6 else '⚠ Low',
         GREEN if aw>=6 else AMBER),
        ('Completion Rate', f"{ac:.1f}%",   '⚠ Low'  if ac<15 else '✅ OK',
         RED if ac<15 else GREEN),
        ('Dest. Clicks',    str(td),
         '❌ Critical' if td==0 else '✅', RED if td==0 else GREEN),
        ('Avg Frequency',   f"{tf:.2f}",    '—', DGREY),
    ]))
    story.append(sp(6))
    kpi2 = [['Total Impressions','Total Reach','Clicks (all)','Avg Frequency']]
    kpi2.append([f"{ti:,}", f"{tr:,}", f"{tall:,}", f"{tf:.2f}"])
    story += [mktbl(kpi2,[W/4]*4,PINK), sp(8)]

    if td==0:
        story.append(P(
            "CRITICAL: Zero destination clicks across all ads — nobody clicked to WhatsApp or idealz.lk. "
            "Fix:  (1) Set a destination URL in every ad.  "
            "(2) Add a Message Us CTA button in ad settings.  "
            "(3) Move CTA text to the first 3 seconds of the video.",
            PS('crit',9,False,RED,sp=8)))

    if prev_df is not None:
        pv_ts=prev_df['Cost'].sum(); pv_tv=int(prev_df['Video views'].sum()); pv_tr=int(prev_df['Reach'].sum())
        pr=[['Metric','This Period','Previous Period','Change']]
        for m,cv,pv in [('Spend (USD)',ts,pv_ts),('Video Views',tv,pv_tv),('Reach',tr,pv_tr)]:
            fv=f"${cv:,.2f}" if 'Spend' in m else f"{int(cv):,}"
            fp=f"${pv:,.2f}" if 'Spend' in m else f"{int(pv):,}"
            pr.append([m,fv,fp,f"{(cv-pv)/pv*100:+.1f}%" if pv>0 else '—'])
        story += [mktbl(pr,[4*cm,4*cm,4*cm,W-12*cm],PINK), sp(10)]

    # Campaign breakdown
    camp = df.groupby('Campaign name').agg(
        Spend=('Cost','sum'), Impressions=('Impressions','sum'), Reach=('Reach','sum'),
        Avg_Freq=('Frequency','mean'), Video_Views=('Video views','sum'),
        Avg_Watch=('Average play time per video view','mean'),
        Avg_Comp=('100% video view rate','mean'),
        Dest_Clicks=('Clicks (destination)','sum'),
    ).round(2).reset_index()
    camp['Comp_%'] = (camp['Avg_Comp']*100).round(1)

    story += [KeepTogether([banner("📂  Campaign Breakdown", PINK), sp(8),
        two_charts(
            bar_chart(camp['Campaign name'].tolist(), camp['Spend'].tolist(),
                      'Spend by Campaign (USD)', color='#D4537E', fmt='${:.2f}',
                      W=Wcm/2-0.5, H=6),
            bar_chart(camp['Campaign name'].tolist(), camp['Avg_Watch'].tolist(),
                      'Avg Watch Time (seconds)', color='#1D9E75', fmt='{:.1f}s',
                      W=Wcm/2-0.5, H=6)),
        sp(4),
        two_charts(
            bar_chart(camp['Campaign name'].tolist(), camp['Comp_%'].tolist(),
                      'Completion Rate (%)', color='#BA7517', fmt='{:.1f}%',
                      W=Wcm/2-0.5, H=6),
            bar_chart(camp['Campaign name'].tolist(), camp['Video_Views'].tolist(),
                      'Video Views', color='#D4537E', W=Wcm/2-0.5, H=6)),
    ]), sp(6)]

    cr_cw = [3.8*cm,2.2*cm,2.5*cm,2.2*cm,1.8*cm,1.8*cm,W-14.3*cm]
    cr = [['Campaign','Spend\n(USD)','Impressions','Video\nViews','Watch\nTime','Comp\n%','Dest\nClicks']]
    for _,r in camp.iterrows():
        cr.append([PW(r['Campaign name'], 17), PC(f"${r['Spend']:,.2f}"),
            PC(f"{int(r['Impressions']):,}"), PC(f"{int(r['Video_Views']):,}"),
            PC(f"{r['Avg_Watch']:.1f}s"), PC(f"{r['Comp_%']:.1f}%"),
            PC(str(int(r['Dest_Clicks'])))])
    story += [mktbl(cr, cr_cw, PINK), sp(12)]

    # Video metrics
    df['comp_%'] = (df['100% video view rate']*100).round(1)
    df['2sec_%'] = (df['2-second video views']/df['Video views'].replace(0,1)*100).round(1)
    df['6sec_%'] = (df['6-second video views']/df['Video views'].replace(0,1)*100).round(1)
    tt_s = df.sort_values('Average play time per video view', ascending=False)
    h_ht = max(7, len(df)*0.55)

    story += [KeepTogether([banner("🎬  Video Metrics Audit", PINK), sp(8),
        two_charts(
            color_bar([shorten(n,18) for n in tt_s['Ad name']],
                      tt_s['Average play time per video view'].tolist(),
                      'Avg Watch Time per Ad (s)', cmap_name='RdYlGn',
                      fmt='{:.1f}s', W=Wcm/2-0.5, H=h_ht),
            color_bar([shorten(n,18) for n in df.sort_values('comp_%',ascending=False)['Ad name']],
                      df.sort_values('comp_%',ascending=False)['comp_%'].tolist(),
                      'Completion Rate per Ad (%)', cmap_name='RdYlGn',
                      fmt='{:.1f}%', W=Wcm/2-0.5, H=h_ht)),
    ]), sp(6)]

    tt2 = df.sort_values('2sec_%',ascending=False)
    c3  = grouped_bar([shorten(n,14) for n in tt2['Ad name']],
                      {'2-sec %':tt2['2sec_%'].tolist(),'6-sec %':tt2['6sec_%'].tolist()},
                      '2-second vs 6-second View Rate per Ad (%)',
                      W=Wcm, H=h_ht, rotate=35)
    story += [c3, sp(6)]

    vm_cw = [3.8*cm,2.8*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,W-12.6*cm]
    vm = [['Ad Name','Campaign','Watch\nTime','Comp\n%','2-sec\n%','6-sec\n%','Dest\nClicks']]
    for _,r in tt_s.iterrows():
        vm.append([PW(r['Ad name'], 17), PW(r['Campaign name'], 12),
            PC(f"{r['Average play time per video view']:.1f}s"), PC(f"{r['comp_%']:.1f}%"),
            PC(f"{r['2sec_%']:.1f}%"), PC(f"{r['6sec_%']:.1f}%"),
            PC(str(int(r['Clicks (destination)'])))])
    story += [mktbl(vm, vm_cw, PINK), sp(4)]
    story.append(P("Benchmark:  0-3s Very weak   3-6s Needs work   6-10s Good   10s+ Excellent   Completion target: 25%+",
                   PS('bm',8,False,AMBER,sp=8)))
    story.append(sp(8))

    # Funnel
    sec2   = int(df['2-second video views'].sum())
    sec6   = int(df['6-second video views'].sum())
    full_v = int((df['Video views']*df['100% video view rate']).sum())
    stages = ['Impressions','Video Views','2-sec Views','6-sec Views','Full Views','Dest. Clicks']
    vals   = [ti, tv, sec2, sec6, full_v, td]
    c_fn   = funnel_chart(stages, vals, 'Video Engagement Funnel', color='#D4537E', W=Wcm, H=9)

    fn_cw = [3*cm,3*cm,3.5*cm,W-9.5*cm]
    fn = [['Stage','Count','% of Impressions','Interpretation']]
    interps=['—','Strong view-through','Initial hook',
             '⚠ Drop-off — move CTA earlier' if sec6<sec2*0.4 else 'Decent retention',
             '❌ Very low' if ac<10 else 'Below average',
             '❌ Fix URL' if td==0 else '✅ OK']
    for s,v,i in zip(stages,vals,interps):
        fn.append([P(s), PC(f"{v:,}"), PC(f"{v/ti*100:.2f}%" if ti>0 else '—'), P(i)])

    story += [KeepTogether([banner("🔻  Video Engagement Funnel", PINK), sp(8),
                             c_fn, sp(6), mktbl(fn, fn_cw, PINK)]), sp(12)]

    # CTR Audit
    ctr_cw = [4.0*cm,3.0*cm,2.0*cm,1.8*cm,2.5*cm,W-13.3*cm]
    ctr = [['Ad Name','Campaign','Dest. CTR','Dest.\nClicks','Impressions','Spend\n(USD)']]
    for _,r in df.sort_values('Cost',ascending=False).iterrows():
        ctr.append([PW(r['Ad name'], 18), PW(r['Campaign name'], 13),
            PC(f"{r['CTR (destination)']:.4f}"), PC(str(int(r['Clicks (destination)']))),
            PC(f"{int(r['Impressions']):,}"), PC(f"${r['Cost']:,.2f}")])
    note = ("All destination CTR values are 0 — nobody clicked to WhatsApp or idealz.lk from any ad."
            if td==0 else "Some ads have zero destination CTR — review those creatives.")
    story += [KeepTogether([
        banner("🔗  Destination CTR Audit", PINK), sp(6),
        P(note, PS('cn2',9,False,RED if td==0 else AMBER,sp=8)),
    ]), mktbl(ctr, ctr_cw, PINK), sp(12)]

    # Monthly benchmarks
    story += [KeepTogether([banner("📈  Monthly Benchmarks", PINK), sp(8),
        kpi_row([('Total Spend',f"${ts:,.2f}",'—',DGREY),
                 ('Impressions',f"{ti:,}",'—',DGREY),
                 ('Total Reach',f"{tr:,}",'—',DGREY),
                 ('Video Views',f"{tv:,}",'—',DGREY),
                 ('Dest. Clicks',str(td),'❌ Critical' if td==0 else '✅',
                  RED if td==0 else GREEN)]),
    ]), sp(16)]

    story += [footer_tbl(analyst, period, "TikTok Ads")]
    doc.build(story)
    buf.seek(0)
    return buf.read()


# Re-export generate_tiktok_pdf pointing to _build_tiktok
def generate_tiktok_pdf(df, analyst, period, alerts=None, prev_df=None):
    return _build_tiktok(df, analyst, period, alerts, prev_df)


# ═════════════════════════════════════════════════════════════════════════════
# COMBINED WEEKLY SUMMARY
# ═════════════════════════════════════════════════════════════════════════════
def generate_pdf_report(meta_df, tiktok_df, ga4_df, analyst, period, recs, alerts, extra=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=1.5*cm, bottomMargin=2*cm)
    story = []
    story += [header_tbl("iDealz Weekly Marketing Report",
                          "Meta, TikTok and Website — Combined Summary",
                          period, analyst), sp(12)]

    al = alert_table(alerts)
    if al: story += [banner("⚠  All Alerts", RED), sp(6), al, sp(12)]

    if meta_df is not None:
        ts=meta_df['Amount spent (LKR)'].sum(); tr=int(meta_df['Results'].sum())
        cpr=ts/tr if tr>0 else 0; ctr=meta_df['CTR (link click-through rate)'].mean()
        story += [banner("📘  Meta Summary", BLUE), sp(8),
                  kpi_row([('Spend',f"Rs {ts:,.0f}",'—',DGREY),
                            ('Results',str(tr),'—',DGREY),
                            ('CPR',f"Rs {cpr:,.2f}",'⚠' if cpr>25 else '✅',RED if cpr>25 else GREEN),
                            ('CTR',f"{ctr:.2f}%",'✅' if ctr>=1.5 else '⚠',GREEN if ctr>=1.5 else AMBER)]),
                  sp(12)]

    if tiktok_df is not None:
        aw=tiktok_df['Average play time per video view'].mean()
        td=int(tiktok_df['Clicks (destination)'].sum())
        story += [banner("🎵  TikTok Summary", PINK), sp(8),
                  kpi_row([('Spend',f"${tiktok_df['Cost'].sum():,.2f}",'—',DGREY),
                            ('Video Views',f"{int(tiktok_df['Video views'].sum()):,}",'—',DGREY),
                            ('Watch Time',f"{aw:.1f}s",'✅' if aw>=6 else '⚠',GREEN if aw>=6 else AMBER),
                            ('Dest. Clicks',str(td),'❌ Critical' if td==0 else '✅',RED if td==0 else GREEN)]),
                  sp(12)]

    if ga4_df is not None:
        ts_w=int(ga4_df['Sessions'].sum()) if 'Sessions' in ga4_df.columns else 0
        er_w=ga4_df['Engagement rate'].mean()*100 if 'Engagement rate' in ga4_df.columns else 0
        story += [banner("🌐  Website Summary", GRNDK), sp(8),
                  kpi_row([('Sessions',f"{ts_w:,}",'—',DGREY),
                            ('Eng. Rate',f"{er_w:.1f}%",'✅' if er_w>50 else '⚠',GREEN if er_w>50 else AMBER),
                            ('Paid Social','Not tracked','❌ UTMs missing',RED)]),
                  sp(12)]

    story += [banner("💡  Recommendations", AMBER), sp(8)]
    rc = [['#','Priority','Recommendation']]
    for i,r in enumerate(recs,1):
        if r.strip():
            p='🔴 Urgent' if i<=2 else ('🟡 High' if i<=4 else '🟢 Normal')
            rc.append([str(i), P(p), PW(r.strip(),80)])
    story += [mktbl(rc,[0.5*cm,2.5*cm,W-3*cm],AMBER), sp(16)]
    story += [footer_tbl(analyst, period, "Weekly Summary")]
    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ═════════════════════════════════════════════════════════════════════════════
def generate_excel_report(meta_df, tiktok_df, ga4_df, period, alerts, extra=None):
    buf = io.BytesIO()
    wb  = Workbook()
    wb.remove(wb.active)

    def hfill(hx): return PatternFill("solid", fgColor=hx.lstrip('#'))
    def tborder():
        s=Side(style='thin',color='FFD3D1C7')
        return Border(left=s,right=s,top=s,bottom=s)
    def write_df(ws, df, hcolor):
        for i,col in enumerate(df.columns,1):
            c=ws.cell(row=1,column=i,value=col)
            c.font=Font(bold=True,color='FFFFFFFF',size=10,name='Arial')
            c.fill=hfill(hcolor)
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.border=tborder()
            ws.column_dimensions[get_column_letter(i)].width=max(14,len(str(col))+4)
        for ri,row in df.iterrows():
            for ci,val in enumerate(row,1):
                c=ws.cell(row=ri+2,column=ci,value=val)
                c.font=Font(size=10,name='Arial')
                c.alignment=Alignment(horizontal='left',vertical='center')
                c.border=tborder()
                c.fill=PatternFill("solid",fgColor='FFFFFFFF' if ri%2==0 else 'FFF9F9F9')
        ws.row_dimensions[1].height=22

    ws=wb.create_sheet("Summary")
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=28
    ws.column_dimensions['C'].width=22; ws.column_dimensions['D'].width=18
    ws.merge_cells('A1:D1'); ws['A1']=f'iDealz Marketing Analytics — {period}'
    ws['A1'].font=Font(bold=True,size=14,color='FFFFFFFF',name='Arial')
    ws['A1'].fill=hfill('#1A1A2E')
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=32

    r=3
    def add_sec(title,rows,color):
        nonlocal r
        ws.merge_cells(f'A{r}:D{r}'); ws[f'A{r}']=title
        ws[f'A{r}'].font=Font(bold=True,size=11,color='FFFFFFFF',name='Arial')
        ws[f'A{r}'].fill=hfill(color)
        ws[f'A{r}'].alignment=Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[r].height=20; r+=1
        for row in rows:
            for ci,val in enumerate(row,1):
                c=ws.cell(row=r,column=ci,value=val)
                c.font=Font(size=10,name='Arial')
                c.fill=PatternFill("solid",fgColor='FFF9F9F9' if r%2 else 'FFFFFFFF')
                c.border=tborder()
            r+=1
        r+=1

    if meta_df is not None:
        ts=meta_df['Amount spent (LKR)'].sum(); tr=int(meta_df['Results'].sum())
        add_sec('META ADS',[
            ['Total Spend (LKR)',f"Rs {ts:,.0f}",'',''],
            ['Total Results',str(tr),'',''],
            ['Cost per Result',f"Rs {ts/tr:,.2f}" if tr>0 else '—','Benchmark: < Rs 25',''],
            ['Avg CTR',f"{meta_df['CTR (link click-through rate)'].mean():.2f}%",'Benchmark: 1.5%+',''],
            ['Avg Frequency',f"{meta_df['Frequency'].mean():.2f}",'Max: 3.0',''],
        ],'#378ADD')

    if tiktok_df is not None:
        td=int(tiktok_df['Clicks (destination)'].sum())
        add_sec('TIKTOK ADS',[
            ['Total Spend (USD)',f"${tiktok_df['Cost'].sum():,.2f}",'',''],
            ['Total Video Views',f"{int(tiktok_df['Video views'].sum()):,}",'',''],
            ['Avg Watch Time',f"{tiktok_df['Average play time per video view'].mean():.1f}s",'Benchmark: 6s+',''],
            ['Completion Rate',f"{tiktok_df['100% video view rate'].mean()*100:.1f}%",'Benchmark: 25%+',''],
            ['Destination Clicks',str(td),'Should be > 0','❌ Critical' if td==0 else '✅'],
        ],'#D4537E')

    if ga4_df is not None:
        nc=ga4_df.select_dtypes(include='number').columns.tolist()
        sc=next((c for c in nc if 'session' in c.lower() and 'engaged' not in c.lower()),nc[0] if nc else None)
        add_sec('WEBSITE (GA4)',[
            ['Total Sessions',f"{int(ga4_df[sc].sum()):,}" if sc else '—','',''],
            ['Paid Social','Not detected — UTMs missing','','❌'],
        ],'#3B6D11')

    if alerts:
        wa=wb.create_sheet("Alerts")
        wa.column_dimensions['A'].width=8; wa.column_dimensions['B'].width=30; wa.column_dimensions['C'].width=60
        for ci,h in enumerate(['Level','Alert','Detail'],1):
            c=wa.cell(row=1,column=ci,value=h)
            c.font=Font(bold=True,color='FFFFFFFF',size=10,name='Arial')
            c.fill=hfill('#E24B4A'); c.border=tborder()
        for ri,a in enumerate(alerts,2):
            wa.cell(row=ri,column=1,value=a['level'].upper()).border=tborder()
            wa.cell(row=ri,column=2,value=a['title']).border=tborder()
            wa.cell(row=ri,column=3,value=a['msg']).border=tborder()

    if meta_df is not None:   write_df(wb.create_sheet("Meta_Raw"),       meta_df,   '#378ADD')
    if tiktok_df is not None: write_df(wb.create_sheet("TikTok_Raw"),     tiktok_df, '#D4537E')
    if ga4_df is not None:    write_df(wb.create_sheet("Website_Channel"),ga4_df,    '#3B6D11')
    if extra:
        for key,sname,color in [('traffic_source','Website_Source','#3B6D11'),
                                  ('users','Website_Users','#3B6D11'),
                                  ('pages','Website_Pages','#27500A'),
                                  ('events','Website_Events','#27500A')]:
            df_ex=extra.get(key)
            if df_ex is not None: write_df(wb.create_sheet(sname),df_ex,color)

    wb.save(buf); buf.seek(0)
    return buf.read()
